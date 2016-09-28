# coding: utf-8

from openpyxl import load_workbook
import re

import db
from datetime_conversion import DateFormat


#  Rewrite in class
def parse(session):  # add date
    wb = load_workbook('1c\dogovori.xlsx')
    ws = wb.active
    rows_num = sum(1 for x in ws.rows)

    for row in range(6, rows_num):
        row_lst = [ws.cell(row=row, column=cell).value for cell in range(1, 18)]
        date_time = DateFormat.erp_str_date(row_lst[0], row_lst[1])
        telephone = '7' + re.sub(r'[\(\)---]', '', row_lst[3])
        contract = db.Erp(date_time, row_lst[2], telephone, row_lst[4], row_lst[5], row_lst[6], row_lst[7], row_lst[8],
                          row_lst[9], row_lst[10], row_lst[11], row_lst[12], row_lst[13], row_lst[14], row_lst[15],
                          row_lst[16])
        session.add(contract)


def merge(session):
    query = session.query(db.Erp.datetime, db.Calltouch.datetime, db.Calltouch.telephone, db.Erp.auto, db.Calltouch.type, db.Calltouch.medium, db.Calltouch.source)\
        .join(db.Calltouch).filter(db.Erp.telephone == db.Calltouch.telephone).group_by(db.Erp.datetime).filter(db.Erp.datetime > db.Calltouch.datetime)
    a = query.all()
    print(len(a))
    for aa in a:
        print(aa)

with db.session_scope() as session:
    # parse(session)
    merge(session)