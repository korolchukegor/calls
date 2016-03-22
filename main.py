# coding: utf-8

import openpyxl

def mod_tel():
    file1 = open('enter.txt')
    res = open('result.txt', 'w')
    for line in file1:
        res.write(line[4:])
    file1.close()

def writefile():

    wb = openpyxl.load_workbook(filename = 'compare.xlsx')
    sheet = wb['test']

    valsA = [sheet.iter_rows('A1:A100')]
    valsB = [sheet.iter_rows('B1:B100')]

    txtfile = open('result.txt')

    j = 1
    #for row in sheet.iter_rows():
    #    for cell in row:
    for line in txtfile:
        sheet.cell('A{}'.format(j)).value = line
        j += 1

    wb.save('compare.xlsx')
writefile()

def compare_tel():

    i = 1
    wb = openpyxl.load_workbook(filename = 'compare.xlsx', use_iterators=True)
    sheet = wb['test']

    valsA = [str(v[0].value) for v in sheet.range('A1:A100')]
    valsB = [str(v[0].value) for v in sheet.range('B1:B100')]

    for valA in valsA:
        for valB in valsB:
            if valA.find(valB) > 0:
                sheet.cell(row = i, column = 3).value = valA
                i += 1
    wb.save('compare.xlsx')


